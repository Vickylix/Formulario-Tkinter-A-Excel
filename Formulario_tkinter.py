import tkinter as tk #--> libreria que me permite usar la interfaz gratis
from tkinter import messagebox #-> importa la caja de mensajes de error que usaremos al no completar la formularia
import re
from openpyxl import Workbook , load_workbook#-> libreria que me crear y manipular una hoja de excel desde python
import os

nombreArchivo =  'datos.xlsx' #-> esto nos ayudara a saber si el archivo ya existe
#comprobar si el archivo ya existe
if os.path.exists(nombreArchivo):
    wb = load_workbook(nombreArchivo)
    ws = wb.active
else :
# creamos una hoja en excel
 wb = Workbook()
 ws = wb.active
 ws.append(["Nombre" , "Apelido" , "Edad" , "Email" , "Telefono" , "Direccion"]) #-> esto crea una lista con los valores que quiero en mi excel

#def para guardar algo xdxd
def guardarDatos():
    nombre = entry_nombre.get()
    apellido = entry_apellido.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    if not nombre or not apellido or not edad or not email or not telefono or not direccion:
        messagebox.showwarning(title="Advertencia", message="Debes completar todos los campos de forma obligatoria.")
        return
    try: # este try confirma si la edad o el telefono son numeros
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning(title="Advertencia", message="La edad y el telefono deben ser numeros")   
        return 
    #en este if validamos que el mail introducido sea un email existente 
    if not re.match(pattern=r"[^@]+@[^@]+\.[^@]+",string=email):
        messagebox.showwarning(title="Advertencia", message="El correo electronico no es valido.")
        return
    ws.append([nombre , apellido , edad , email , telefono , direccion])
    wb.save('nombreArchivo') #-> crea el excel con los datos antes introducidos
    messagebox.showinfo(title="Informacion" , message="Los Datos guardados con exito.")

#limpiar el formulario una vez que se guarda los datos.
    entry_nombre.delete(first=0, last=tk.END)
    entry_apellido.delete(first=0,last=tk.END)
    entry_edad.delete(first=0,last=tk.END)
    entry_email.delete(first=0, last=tk.END)
    entry_telefono.delete(first=0,last=tk.END)
    entry_direccion.delete(first=0,last=tk.END)

## linea 52 hasta la 87 unicamente a la interfaz , letras y color del formulario
root = tk. Tk()
root.title("formulario de base de datos")
root.configure(bg="#1E1E2E")
label_style = {"bg": "#1E1E2E", "fg": "#FFFFFF"}
entry_style = {"bg": "#313244", "fg": "white", "insertbackground": "white"}
## fila del nombre
label_nombre = tk.Label(root, text="Nombre" , **label_style)
label_nombre.grid(row=0, column=0, padx=15, pady=8)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=15, pady=8)
## fila del apellido
label_apellido = tk.Label(root, text="Apellido" , **label_style)
label_apellido.grid(row=1, column=0, padx=15, pady=8)
entry_apellido = tk.Entry(root, **entry_style)
entry_apellido.grid(row=1, column=1, padx=15, pady=8)
## fila de la edad
label_edad = tk.Label(root, text="Edad" , **label_style)
label_edad.grid(row=2, column=0, padx=15, pady=8)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=2, column=1, padx=15, pady=8)
## fila del email
label_email = tk.Label(root, text="Email" , **label_style)
label_email.grid(row=3, column=0, padx=15, pady=8)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=3, column=1, padx=15, pady=8)
##fila del telefono
label_telefono = tk.Label(root, text="Telefono" , **label_style)
label_telefono.grid(row=4, column=0, padx=15, pady=8)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=4, column=1, padx=15, pady=8)
##fila de la direccion
label_direccion = tk.Label(root, text="Direccion" , **label_style)
label_direccion.grid(row=5, column=0, padx=15, pady=8)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=5, column=1, padx=15, pady=8)
##boton para guardar los datos
boton_guardar = tk.Button(
    root,
    text="Guardar",
    command=guardarDatos,
    bg="#89B4FA",
    fg="black",
    activebackground="#74C7EC",
    relief="flat"
)
boton_guardar.grid(row=6,column=0, columnspan=2, padx=10, pady=10)
##
root.mainloop()